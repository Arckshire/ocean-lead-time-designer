# =====================================================================
# Lead Time Optimization — v7 Streamlit App
#
# Architecture:
#   - All existing computation logic preserved verbatim from your app.py
#     (compute_shipment_leadtimes, build_carrier_lane_report,
#      compute_insights_for_metric, write_excel_*, etc.)
#   - 5 new DERIVED helpers added (reliability, skew, trend, bottleneck,
#     dynamic-scope problematic) — all operate on existing outputs.
#   - UI built with heavy custom CSS to match the v7 hybrid mockup.
#   - File-upload landing → transitions to full v7 UI once a file is fed.
#
# Deploy:
#   pip install -r requirements.txt
#   streamlit run app.py
# =====================================================================

import io
import math
from typing import Dict, Optional, List, Tuple

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import streamlit as st


# =====================================================================
# CONSTANTS (verbatim from your app.py)
# =====================================================================
MILESTONES = ["CEP", "CGI", "CLL", "VDL", "VAD", "CDD", "CGO", "CER"]
ORDERED_MILESTONES = ["CEP", "CGI", "CLL", "VDL", "VAD", "CDD", "CGO", "CER"]
SEGMENTS = [
    ("CEP", "CGI"), ("CGI", "CLL"), ("CLL", "VDL"),
    ("VDL", "VAD"), ("VAD", "CDD"), ("CDD", "CGO"), ("CGO", "CER"),
]
P44_FALLBACKS = {"VDL": "VDL_P44", "VAD": "VAD_P44"}
REQUIRED_BASE_COLS = [
    "TENANT_NAME", "MASTER_SHIPMENT_ID", "POL", "POD",
    "CARRIER_NAME", "CARRIER_SCAC",
]
MILESTONE_LONG = {
    "CEP": ("Container Empty Pickup", "POL"),
    "CGI": ("Container Gate-In", "POL"),
    "CLL": ("Container Loaded onto Vessel", "POL"),
    "VDL": ("Vessel Departure from POL", "POL"),
    "VAD": ("Vessel Arrival at POD", "POD"),
    "CDD": ("Container Discharge at POD", "POD"),
    "CGO": ("Container Gate Out at POD", "POD"),
    "CER": ("Container Empty Return", "POD"),
}
DISPLAY_COLS = {
    "TENANT_NAME": "Tenant Name", "LANE": "Lane",
    "CARRIER_NAME": "Carrier Name", "CARRIER_SCAC": "Carrier SCAC",
    "VOLUME": "Volume (Shipments)",
    "TOTAL_H": "Total Lead Time (Hours)", "TOTAL_D": "Total Lead Time (Days)",
    "MIN_H": "Min Lead Time (Hours)", "MIN_D": "Min Lead Time (Days)",
    "MED_H": "Median Lead Time (Hours)", "MED_D": "Median Lead Time (Days)",
    "PCT_H": "P{p} Lead Time (Hours)", "PCT_D": "P{p} Lead Time (Days)",
    "MAX_H": "Max Lead Time (Hours)", "MAX_D": "Max Lead Time (Days)",
}


# =====================================================================
# CSS — v7 visual identity
# =====================================================================
V7_CSS = """
<style>
  :root {
    --bg: #0a0e1a; --bg-elev: #0e1320; --bg-card: #151a2e;
    --bg-card-hover: #1d243e; --bg-card-active: #232a47;
    --border: #252b40; --border-strong: #313a59;
    --text: #ffffff; --text-secondary: #c9d0e0;
    --text-tertiary: #9ca3af; --text-quaternary: #6b7280;
    --accent: #6366f1; --accent-light: #818cf8;
    --accent-glow: rgba(99, 102, 241, 0.18);
    --pink: #ec4899; --pink-glow: rgba(236, 72, 153, 0.18);
    --green: #10b981; --green-glow: rgba(16, 185, 129, 0.15);
    --orange: #f97316; --orange-glow: rgba(249, 115, 22, 0.15);
    --red: #ef4444; --red-glow: rgba(239, 68, 68, 0.15);
    --blue: #3b82f6; --yellow: #fbbf24;
  }
  /* Streamlit overrides */
  .stApp { background: var(--bg); color: var(--text); }
  .block-container { padding-top: 1rem !important; padding-bottom: 2rem !important; max-width: 1600px !important; }
  header[data-testid="stHeader"] { background: transparent; }
  div[data-testid="stToolbar"] { display: none; }
  .stMarkdown, .stMarkdown p { color: var(--text); }
  /* Streamlit buttons styled as v7 buttons */
  div.stButton > button {
    background: var(--bg-card); border: 1px solid var(--border) !important;
    color: var(--text) !important; padding: 8px 14px !important; border-radius: 8px !important;
    font-size: 13px !important; font-weight: 500 !important; transition: all 0.15s;
  }
  div.stButton > button:hover { background: var(--bg-card-hover); border-color: var(--border-strong) !important; }
  div.stButton > button[kind="primary"] { background: var(--accent) !important; border-color: var(--accent) !important; }
  /* Streamlit selectbox / number_input */
  div[data-baseweb="select"] > div { background: var(--bg-elev) !important; border-color: var(--border) !important; }
  div[data-baseweb="select"] * { color: var(--text) !important; }
  div[data-testid="stNumberInput"] input { background: var(--bg-elev) !important; color: var(--text) !important; border-color: var(--border) !important; }
  div[data-testid="stTextInput"] input { background: var(--bg-elev) !important; color: var(--text) !important; border-color: var(--border) !important; }
  /* Streamlit radio segmented */
  div[role="radiogroup"] { gap: 4px !important; background: var(--bg-elev); border: 1px solid var(--border); border-radius: 7px; padding: 3px; display: inline-flex !important; flex-wrap: nowrap !important; }
  div[role="radiogroup"] label { background: transparent !important; padding: 5px 12px !important; border-radius: 5px !important; margin: 0 !important; cursor: pointer; }
  div[role="radiogroup"] label:has(input:checked) { background: var(--bg-card-active) !important; color: var(--text) !important; }
  div[role="radiogroup"] label > div:first-child { display: none !important; }
  /* Streamlit file_uploader */
  div[data-testid="stFileUploaderDropzone"] {
    background: var(--bg-card) !important; border: 2px dashed var(--border-strong) !important;
    border-radius: 14px !important; padding: 32px !important;
  }
  div[data-testid="stFileUploaderDropzone"]:hover { border-color: var(--accent) !important; background: var(--bg-card-hover) !important; }
  /* Hide Streamlit footer + menu */
  footer { visibility: hidden; }
  #MainMenu { visibility: hidden; }

  /* ============================
     v7 CUSTOM COMPONENTS
     ============================ */
  .v7-breadcrumb { display: flex; gap: 8px; color: var(--text-quaternary); font-size: 12.5px; margin-bottom: 4px; }
  .v7-breadcrumb .sep { opacity: 0.4; } .v7-breadcrumb .current { color: var(--text); font-weight: 600; }
  .v7-page-title { font-size: 22px; font-weight: 700; letter-spacing: -0.02em; color: var(--text); margin-bottom: 4px; }
  .v7-date-range { background: var(--bg-card); border: 1px solid var(--border); padding: 8px 14px; border-radius: 8px; font-size: 13px; color: var(--text-secondary); display: inline-flex; gap: 8px; align-items: center; }
  .v7-date-range strong { color: var(--text); font-weight: 500; }

  /* Upload landing */
  .v7-upload-hero { text-align: center; padding: 60px 20px 40px; max-width: 720px; margin: 0 auto; }
  .v7-upload-hero .logo { width: 56px; height: 56px; border-radius: 14px; background: linear-gradient(135deg, var(--accent), var(--pink)); display: inline-grid; place-items: center; font-weight: 700; color: white; font-size: 22px; margin-bottom: 20px; }
  .v7-upload-hero h1 { font-size: 32px; font-weight: 800; letter-spacing: -0.025em; color: var(--text); margin-bottom: 10px; }
  .v7-upload-hero p { font-size: 15px; color: var(--text-secondary); line-height: 1.6; margin-bottom: 22px; }
  .v7-upload-hero .tag { display: inline-block; padding: 4px 12px; border-radius: 12px; background: rgba(99, 102, 241, 0.15); color: var(--accent-light); font-size: 11px; font-weight: 700; letter-spacing: 0.1em; text-transform: uppercase; margin-bottom: 18px; }

  /* Configure strip */
  .v7-configure-strip { background: linear-gradient(180deg, var(--bg-card) 0%, var(--bg-elev) 100%); border: 1px solid var(--border); border-radius: 12px; padding: 12px 18px; margin-bottom: 18px; }
  .v7-cf-row { display: flex; flex-wrap: wrap; gap: 18px; align-items: center; }
  .v7-cf-block { display: flex; flex-direction: column; gap: 2px; }
  .v7-cf-block .l { font-size: 10px; color: var(--text-quaternary); text-transform: uppercase; letter-spacing: 0.06em; font-weight: 600; }
  .v7-cf-block .v { font-size: 13.5px; font-weight: 600; color: var(--text); }
  .v7-cf-block .v .code { color: var(--accent-light); font-family: 'SF Mono', monospace; font-weight: 700; }
  .v7-cf-sep { width: 1px; height: 30px; background: var(--border); }

  /* KPI strip */
  .v7-kpi-card { background: var(--bg-card); border: 1px solid var(--border); border-radius: 12px; padding: 16px 18px; position: relative; overflow: hidden; }
  .v7-kpi-card::after { content: ''; position: absolute; top: 0; right: 0; width: 60px; height: 60px; background: radial-gradient(circle at top right, var(--accent-glow), transparent 70%); pointer-events: none; }
  .v7-kpi-card .label { font-size: 11px; color: var(--text-tertiary); text-transform: uppercase; letter-spacing: 0.05em; font-weight: 600; margin-bottom: 8px; }
  .v7-kpi-card .value { font-size: 26px; font-weight: 700; letter-spacing: -0.02em; color: var(--text); }
  .v7-kpi-card .sub { font-size: 11.5px; color: var(--text-tertiary); margin-top: 6px; }
  .v7-mini-bar { height: 4px; background: var(--border); border-radius: 2px; margin-top: 8px; overflow: hidden; }
  .v7-mini-bar-fill { height: 100%; background: linear-gradient(90deg, var(--accent), var(--pink)); }

  /* Journey hero card */
  .v7-journey-card { background: linear-gradient(180deg, var(--bg-card) 0%, var(--bg-elev) 100%); border: 1px solid var(--border); border-radius: 14px; padding: 22px 26px; margin-bottom: 22px; }
  .v7-jh-eyebrow { font-size: 10.5px; font-weight: 700; letter-spacing: 0.14em; text-transform: uppercase; color: var(--accent-light); margin-bottom: 6px; }
  .v7-jh-title { font-size: 22px; font-weight: 700; letter-spacing: -0.025em; color: var(--text); margin-bottom: 4px; }
  .v7-jh-subtitle { font-size: 12.5px; color: var(--text-tertiary); line-height: 1.5; max-width: 720px; }

  /* Journey diagram */
  .v7-journey-diagram { background: var(--bg-elev); border-radius: 12px; padding: 24px 20px 18px; margin: 14px 0; }
  .v7-jd-port-headers { display: grid; grid-template-columns: 4fr 1fr 4fr; gap: 24px; margin-bottom: 14px; font-size: 10.5px; font-weight: 700; letter-spacing: 0.1em; text-transform: uppercase; }
  .v7-jd-port-headers .pol { color: var(--accent-light); }
  .v7-jd-port-headers .transit { color: var(--text-tertiary); text-align: center; }
  .v7-jd-port-headers .pod { color: var(--pink); text-align: right; }
  .v7-jd-flow { position: relative; display: grid; grid-template-columns: repeat(8, 1fr); gap: 8px; padding: 20px 0 12px; }
  .v7-jd-track-bg { position: absolute; top: 50%; left: 6%; right: 6%; height: 4px; background: var(--border); border-radius: 2px; z-index: 0; transform: translateY(-50%); }
  .v7-jd-track-active { position: absolute; top: 50%; height: 4px; background: linear-gradient(90deg, var(--accent), var(--pink)); z-index: 1; border-radius: 2px; transform: translateY(-50%); box-shadow: 0 0 16px var(--accent-glow); }
  .v7-jd-milestone { position: relative; z-index: 3; display: flex; flex-direction: column; align-items: center; }
  .v7-jd-milestone .dot { width: 36px; height: 36px; border-radius: 50%; background: var(--bg-card); border: 2px solid var(--border); display: grid; place-items: center; font-size: 12px; font-weight: 700; color: var(--text-secondary); }
  .v7-jd-milestone.endpoint .dot { background: var(--accent); border-color: var(--accent); color: white; box-shadow: 0 0 0 5px var(--accent-glow); }
  .v7-jd-milestone.in-path .dot { background: var(--bg-card-active); border-color: var(--accent); color: var(--accent-light); }
  .v7-jd-milestone .code { margin-top: 10px; font-size: 11.5px; font-weight: 700; color: var(--text); }
  .v7-jd-milestone .label { font-size: 10px; color: var(--text-tertiary); text-align: center; margin-top: 2px; line-height: 1.3; }
  .v7-jd-milestone .place { font-size: 9px; font-weight: 700; letter-spacing: 0.06em; text-transform: uppercase; margin-top: 6px; padding: 2px 6px; border-radius: 6px; }
  .v7-jd-milestone .place.pol { background: rgba(99, 102, 241, 0.15); color: var(--accent-light); }
  .v7-jd-milestone .place.pod { background: rgba(236, 72, 153, 0.15); color: var(--pink); }

  /* Whole Journey aggregate card */
  .v7-jd-whole-journey { background: linear-gradient(135deg, rgba(99, 102, 241, 0.22), rgba(236, 72, 153, 0.1), var(--bg-card)); border: 1px solid var(--accent); box-shadow: 0 0 0 1px var(--accent), 0 4px 16px var(--accent-glow); border-radius: 12px; padding: 14px 18px; margin: 16px 6% 10px; display: grid; grid-template-columns: auto 1fr auto auto auto auto auto; gap: 24px; align-items: center; }
  .v7-jd-whole-journey.inactive { background: linear-gradient(135deg, rgba(99, 102, 241, 0.12), rgba(236, 72, 153, 0.06), var(--bg-card)); border-color: var(--border-strong); box-shadow: none; }
  .v7-jd-whole-journey .wj-icon { width: 36px; height: 36px; border-radius: 50%; background: linear-gradient(135deg, var(--accent), var(--pink)); display: grid; place-items: center; font-size: 16px; color: white; }
  .v7-jd-whole-journey .wj-label .l { font-size: 10px; font-weight: 700; letter-spacing: 0.1em; text-transform: uppercase; color: var(--accent-light); margin-bottom: 2px; }
  .v7-jd-whole-journey .wj-label .v { font-size: 14px; font-weight: 700; color: var(--text); }
  .v7-jd-whole-journey .wj-stat { display: flex; flex-direction: column; gap: 2px; text-align: center; }
  .v7-jd-whole-journey .wj-stat .l { font-size: 9.5px; color: var(--text-tertiary); text-transform: uppercase; letter-spacing: 0.06em; }
  .v7-jd-whole-journey .wj-stat .v { font-size: 16px; font-weight: 700; letter-spacing: -0.01em; color: var(--text); }
  .v7-jd-whole-journey .wj-stat .v .u { font-size: 10px; color: var(--text-tertiary); font-weight: 400; }
  .v7-jd-whole-journey .wj-selected-pill { background: var(--accent); color: white; font-size: 10px; font-weight: 700; padding: 4px 10px; border-radius: 12px; letter-spacing: 0.05em; }

  /* Segment cards (rendered as Streamlit buttons styled as cards) */
  .v7-segment-grid-wrap { padding: 0 6%; margin-top: 6px; }
  div[data-testid="stHorizontalBlock"] div.stButton > button[key^="seg_"] {
    width: 100% !important; min-height: 80px !important;
    background: var(--bg-card) !important; border: 1px solid var(--border) !important;
    border-radius: 10px !important; padding: 10px 12px !important;
    font-size: 11px !important; color: var(--text-secondary) !important;
    text-align: left !important; white-space: pre-line !important; line-height: 1.4 !important;
  }
  div[data-testid="stHorizontalBlock"] div.stButton > button[key^="seg_active_"] {
    background: linear-gradient(135deg, rgba(99,102,241,0.12), rgba(236,72,153,0.08)) !important;
    border-color: var(--accent) !important;
    box-shadow: 0 0 0 1px var(--accent), 0 4px 14px var(--accent-glow) !important;
  }
  div[data-testid="stHorizontalBlock"] div.stButton > button[key^="seg_bottleneck_"] {
    border-color: var(--red) !important;
    background: linear-gradient(135deg, rgba(239,68,68,0.05), var(--bg-card)) !important;
  }

  /* Active Scope summary bar */
  .v7-jh-summary { display: grid; grid-template-columns: 1.4fr 1fr 1fr 1fr 1.2fr; gap: 14px; background: var(--bg-elev); border-radius: 10px; padding: 14px 18px; margin-top: 6px; }
  .v7-jws-block { display: flex; flex-direction: column; gap: 2px; }
  .v7-jws-block .l { font-size: 10px; color: var(--text-tertiary); text-transform: uppercase; letter-spacing: 0.06em; }
  .v7-jws-block .v { font-size: 17px; font-weight: 700; letter-spacing: -0.02em; color: var(--text); }
  .v7-jws-block .v .u { font-size: 11px; color: var(--text-tertiary); font-weight: 500; margin-left: 2px; }

  /* Bottleneck callout */
  .v7-bottleneck-callout { background: linear-gradient(135deg, rgba(239,68,68,0.10), rgba(249,115,22,0.05)); border: 1px solid rgba(239,68,68,0.4); border-radius: 14px; padding: 16px 22px; margin-bottom: 22px; display: grid; grid-template-columns: auto 1fr; gap: 18px; align-items: center; }
  .v7-bc-icon { font-size: 24px; }
  .v7-bc-text h3 { font-size: 15px; font-weight: 700; color: var(--text); margin-bottom: 4px; }
  .v7-bc-text h3 strong { color: var(--red); }
  .v7-bc-text p { font-size: 12.5px; color: var(--text-secondary); line-height: 1.55; }

  /* Lane context */
  .v7-lane-context { padding: 12px 14px; background: var(--bg-card); border: 1px solid var(--border); border-radius: 10px; margin-bottom: 18px; }
  .v7-lane-context .lc-label { font-size: 12px; color: var(--text-tertiary); font-weight: 600; }
  .v7-lane-context .lc-stats { display: inline-flex; gap: 18px; font-size: 12px; color: var(--text-secondary); margin-left: 18px; }
  .v7-lane-context .lc-stats strong { color: var(--text); font-weight: 600; }

  /* Section heads */
  .v7-section-head { display: flex; justify-content: space-between; align-items: flex-end; margin: 26px 0 14px; }
  .v7-section-eyebrow { font-size: 10px; font-weight: 700; letter-spacing: 0.14em; text-transform: uppercase; color: var(--text-quaternary); margin-bottom: 4px; }
  .v7-section-title { font-size: 17px; font-weight: 700; letter-spacing: -0.015em; color: var(--text); }
  .v7-section-subtitle { font-size: 12px; color: var(--text-tertiary); margin-top: 2px; }

  /* Insight cards */
  .v7-insight-card { background: var(--bg-card); border: 1px solid var(--border); border-radius: 12px; padding: 16px 18px; height: 100%; }
  .v7-insight-card.featured { background: linear-gradient(135deg, var(--bg-card), rgba(99,102,241,0.04)); border-color: rgba(99,102,241,0.45); }
  .v7-ic-head { display: flex; justify-content: space-between; align-items: center; margin-bottom: 12px; gap: 10px; }
  .v7-ic-title { font-size: 12px; font-weight: 600; color: var(--text-tertiary); text-transform: uppercase; letter-spacing: 0.06em; display: flex; align-items: center; gap: 6px; }
  .v7-tag-ai { background: var(--pink-glow); color: var(--pink); padding: 2px 6px; border-radius: 8px; font-size: 9px; font-weight: 700; letter-spacing: 0.06em; }

  /* Best Carrier */
  .v7-bc-rank { display: inline-flex; align-items: center; gap: 6px; background: linear-gradient(135deg, #fbbf24, #f59e0b); color: #1a1300; padding: 3px 10px; border-radius: 12px; font-size: 11px; font-weight: 700; margin-bottom: 12px; }
  .v7-bc-name { font-size: 22px; font-weight: 700; letter-spacing: -0.02em; color: var(--text); margin-bottom: 4px; }
  .v7-bc-scac { font-size: 11px; color: var(--text-quaternary); font-family: 'SF Mono', monospace; margin-bottom: 14px; }
  .v7-bc-quote { font-size: 13px; color: var(--text); line-height: 1.6; margin-bottom: 14px; padding: 12px 14px; background: var(--bg-elev); border-left: 3px solid var(--accent); border-radius: 6px; }
  .v7-bc-quote strong { color: var(--accent-light); font-weight: 600; }
  .v7-bc-stats { display: grid; grid-template-columns: 1fr 1fr; gap: 10px; }
  .v7-bc-stat { background: var(--bg-elev); border-radius: 8px; padding: 8px 10px; }
  .v7-bc-stat .l { font-size: 10px; color: var(--text-quaternary); text-transform: uppercase; }
  .v7-bc-stat .v { font-size: 15px; font-weight: 600; color: var(--text); margin-top: 2px; }

  /* Top 5 list */
  .v7-rec-list { display: flex; flex-direction: column; gap: 8px; }
  .v7-rec-row { display: grid; grid-template-columns: 28px 1fr auto; gap: 12px; align-items: center; padding: 10px 12px; background: var(--bg-elev); border-radius: 8px; }
  .v7-rec-row.rank-1 { background: linear-gradient(90deg, rgba(251, 191, 36, 0.08), var(--bg-elev)); }
  .v7-rec-row .rank { width: 26px; height: 26px; border-radius: 50%; background: var(--bg-card); display: grid; place-items: center; font-size: 12px; font-weight: 700; color: var(--text-secondary); }
  .v7-rec-row.rank-1 .rank { background: linear-gradient(135deg, #fbbf24, #f59e0b); color: #1a1300; }
  .v7-rec-row .info .name { font-size: 13px; font-weight: 600; color: var(--text); }
  .v7-rec-row .info .why { font-size: 11px; color: var(--text-tertiary); margin-top: 2px; }
  .v7-rec-row .stats { display: flex; gap: 12px; }
  .v7-rec-row .stats .stat { text-align: right; }
  .v7-rec-row .stats .stat .v { font-size: 13px; font-weight: 600; color: var(--text); }
  .v7-rec-row .stats .stat .l { font-size: 9px; color: var(--text-tertiary); text-transform: uppercase; }

  /* Problematic Shipments */
  .v7-prob-scope { padding: 8px 10px; background: var(--bg-elev); border-radius: 6px; margin-bottom: 12px; border-left: 3px solid var(--accent); }
  .v7-prob-scope .l { font-size: 9.5px; font-weight: 700; letter-spacing: 0.08em; text-transform: uppercase; color: var(--text-tertiary); }
  .v7-prob-scope .v { font-size: 12.5px; font-weight: 600; color: var(--text); margin-top: 1px; }
  .v7-prob-list { display: flex; flex-direction: column; gap: 6px; }
  .v7-prob-row { display: grid; grid-template-columns: 1fr auto auto; gap: 12px; align-items: center; padding: 9px 12px; background: var(--bg-elev); border-radius: 7px; border-left: 3px solid var(--red); font-size: 12px; }
  .v7-prob-row.med { border-left-color: var(--orange); }
  .v7-prob-row .shipment-id { font-family: 'SF Mono', monospace; font-size: 11px; color: var(--text-tertiary); }
  .v7-prob-row .shipment-id .c { color: var(--text); font-weight: 600; display: block; }
  .v7-prob-row .actual { font-weight: 600; font-size: 13px; color: var(--text); }
  .v7-prob-row .delta { color: var(--red); font-size: 11.5px; font-weight: 600; }
  .v7-prob-row.med .delta { color: var(--orange); }
  .v7-prob-footer { margin-top: 12px; padding-top: 12px; border-top: 1px solid var(--border); font-size: 11.5px; color: var(--text-tertiary); }
  .v7-prob-footer strong { color: var(--text); }

  /* Skew rows */
  .v7-skew-row { display: grid; grid-template-columns: 84px 1fr 92px; gap: 12px; align-items: center; padding: 10px 0; border-bottom: 1px solid var(--border); }
  .v7-skew-row:last-child { border-bottom: none; }
  .v7-skew-row .what { font-size: 12px; color: var(--text-secondary); }
  .v7-skew-row .what .count { font-size: 10.5px; color: var(--text-quaternary); display: block; }
  .v7-skew-row .bar { height: 6px; background: var(--bg-elev); border-radius: 3px; overflow: hidden; }
  .v7-skew-row .bar .fill { height: 100%; background: linear-gradient(90deg, var(--accent), var(--pink)); }
  .v7-skew-row .result { font-size: 12.5px; font-weight: 600; text-align: right; color: var(--green); }
  .v7-skew-row .result .from { color: var(--text-quaternary); font-weight: 400; font-size: 11px; display: block; }

  /* Reliability */
  .v7-reliability-explanation { margin-top: 14px; padding: 12px 14px; background: var(--bg-elev); border-radius: 8px; border-left: 3px solid var(--accent-light); }
  .v7-re-head { font-size: 11px; font-weight: 700; color: var(--accent-light); letter-spacing: 0.06em; text-transform: uppercase; margin-bottom: 8px; }
  .v7-re-body { font-size: 12.5px; color: var(--text-secondary); line-height: 1.65; }
  .v7-re-body strong { color: var(--text); }
  .v7-re-body ul { list-style: none; padding-left: 0; margin: 6px 0; }
  .v7-re-body ul li { padding: 4px 0; padding-left: 22px; position: relative; }
  .v7-re-body ul li::before { content: '✓'; position: absolute; left: 0; color: var(--green); font-weight: 700; }

  /* Concentration */
  .v7-conc-flag { margin-top: 12px; padding: 8px 12px; background: var(--orange-glow); border-radius: 6px; color: var(--orange); font-size: 11.5px; display: flex; align-items: center; gap: 8px; }
  .v7-conc-flag.low { background: var(--green-glow); color: var(--green); }
  .v7-conc-flag.high { background: var(--red-glow); color: var(--red); }
  .v7-conc-flag strong { font-weight: 700; }

  /* Analyst table */
  .v7-analyst-table { background: var(--bg-card); border: 1px solid var(--border); border-radius: 12px; overflow: hidden; }
  .v7-analyst-table table { width: 100%; border-collapse: collapse; font-size: 12.5px; }
  .v7-analyst-table thead th { background: var(--bg-elev); padding: 12px 14px; text-align: left; font-size: 10.5px; font-weight: 600; color: var(--text-tertiary); text-transform: uppercase; letter-spacing: 0.05em; border-bottom: 1px solid var(--border); }
  .v7-analyst-table tbody td { padding: 11px 14px; border-bottom: 1px solid var(--border); color: var(--text); }
  .v7-analyst-table tbody tr.lane-row { background: var(--bg-elev); }
  .v7-analyst-table tbody tr.lane-row td { font-weight: 600; font-size: 13px; }
  .v7-analyst-table tbody tr.lane-row td:first-child { color: var(--accent-light); }
  .v7-analyst-table tbody tr.carrier-row td:first-child { padding-left: 32px; color: var(--text-secondary); }
  .v7-pill { display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 11px; font-weight: 500; }
  .v7-pill.green { background: var(--green-glow); color: var(--green); }
  .v7-pill.gray { background: var(--bg-card-active); color: var(--text-secondary); }
</style>
"""


# =====================================================================
# EXISTING FUNCTIONS — preserved VERBATIM from your app.py
# (only the section comments are slightly trimmed for compactness)
# =====================================================================
def _read_input(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        encodings = ["utf-8", "utf-8-sig", "cp1252", "latin-1"]
        last_err = None
        for enc in encodings:
            try:
                uploaded_file.seek(0)
                return pd.read_csv(uploaded_file, encoding=enc)
            except UnicodeDecodeError as e:
                last_err = e
                continue
        raise ValueError(f"Unable to decode CSV with {encodings}. Last error: {last_err}.")
    if name.endswith(".xlsx") or name.endswith(".xls"):
        uploaded_file.seek(0)
        return pd.read_excel(uploaded_file)
    raise ValueError("Unsupported file type. Please upload a CSV or Excel file.")


def get_missing_columns(df, start_ms, end_ms, whole_journey):
    missing = []
    for col in REQUIRED_BASE_COLS:
        if col not in df.columns:
            missing.append(col)
    selected_milestones = {start_ms, end_ms}
    if whole_journey:
        selected_milestones.update(ORDERED_MILESTONES)
    for ms in sorted(selected_milestones):
        if ms in P44_FALLBACKS:
            fb = P44_FALLBACKS[ms]
            if ms not in df.columns and fb not in df.columns:
                missing.append(f"{ms} (or {fb})")
        else:
            if ms not in df.columns:
                missing.append(ms)
    return missing


def _coerce_datetimes(df, cols):
    for c in cols:
        if c in df.columns:
            s = pd.to_datetime(df[c], errors="coerce", utc=True)
            df[c] = s.dt.tz_convert(None)
    return df


def _resolve_timestamp(row, milestone):
    val = row.get(milestone, pd.NaT)
    if pd.notna(val):
        return val
    fb = P44_FALLBACKS.get(milestone)
    if fb:
        return row.get(fb, pd.NaT)
    return pd.NaT


def _round_hours(x):
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return None
    return float(np.round(x, 2))


def _round_days_from_hours(x_hours):
    if x_hours is None or (isinstance(x_hours, float) and np.isnan(x_hours)):
        return None
    return int(np.round(x_hours / 24.0))


def _safe_quantile(values, q):
    values = values.dropna()
    if values.empty:
        return None
    return float(values.quantile(q, interpolation="linear"))


def _make_lane(pol, pod):
    pol_s = "" if pd.isna(pol) else str(pol)
    pod_s = "" if pd.isna(pod) else str(pod)
    return f"{pol_s} → {pod_s}"


def _segment_col_name(start_ms, end_ms):
    return f"SEG_{start_ms}_{end_ms}_HOURS"


def _journey_col_name():
    return "JOURNEY_LEAD_HOURS"


def _pct_to_count(total, pct):
    if total <= 0 or pct <= 0:
        return 0
    return int(np.ceil(total * (pct / 100.0)))


def add_shipment_month_year(raw_df):
    df = raw_df.copy()
    dt_cols = ORDERED_MILESTONES + list(P44_FALLBACKS.values())
    dt_cols = list(dict.fromkeys(dt_cols))
    df = _coerce_datetimes(df, dt_cols)
    for ms in ORDERED_MILESTONES:
        df[f"_RES_{ms}"] = df.apply(lambda r: _resolve_timestamp(r, ms), axis=1)
        df[f"_RES_{ms}"] = pd.to_datetime(df[f"_RES_{ms}"], errors="coerce", utc=True).dt.tz_convert(None)
    first_ts = pd.Series(pd.NaT, index=df.index, dtype="datetime64[ns]")
    for ms in ORDERED_MILESTONES:
        first_ts = first_ts.fillna(df[f"_RES_{ms}"])
    df["SHIPMENT_MONTH_YEAR"] = first_ts.dt.strftime("%b %Y").fillna("")
    df["_FIRST_TS"] = first_ts
    df = df.drop(columns=[f"_RES_{ms}" for ms in ORDERED_MILESTONES], errors="ignore")
    return df


def compute_shipment_leadtimes(raw, start_ms, end_ms, shipment_agg, whole_journey):
    df = raw.copy()
    dt_cols = [start_ms, end_ms] + ORDERED_MILESTONES + list(P44_FALLBACKS.values())
    dt_cols = list(dict.fromkeys(dt_cols))
    df = _coerce_datetimes(df, dt_cols)
    for ms in ORDERED_MILESTONES:
        df[f"_RES_{ms}"] = df.apply(lambda r: _resolve_timestamp(r, ms), axis=1)
        df[f"_RES_{ms}"] = pd.to_datetime(df[f"_RES_{ms}"], errors="coerce", utc=True).dt.tz_convert(None)
    df["_START_TS"] = df[f"_RES_{start_ms}"]
    df["_END_TS"] = df[f"_RES_{end_ms}"]
    group_cols = ["TENANT_NAME", "POL", "POD", "CARRIER_NAME", "CARRIER_SCAC", "MASTER_SHIPMENT_ID"]
    if whole_journey:
        all_present = pd.Series(True, index=df.index)
        for ms in ORDERED_MILESTONES:
            all_present = all_present & pd.notna(df[f"_RES_{ms}"])
        ordered = pd.Series(True, index=df.index)
        for a, b in SEGMENTS:
            ordered = ordered & (df[f"_RES_{b}"] >= df[f"_RES_{a}"])
        selected = pd.notna(df["_START_TS"]) & pd.notna(df["_END_TS"]) & (df["_END_TS"] >= df["_START_TS"])
        df["_WHOLE_VALID"] = all_present & ordered & selected
        qdf = df[df["_WHOLE_VALID"]].copy()
        if qdf.empty:
            return pd.DataFrame(columns=group_cols + ["LANE", _journey_col_name()] + [_segment_col_name(a, b) for a, b in SEGMENTS])
        qdf[_journey_col_name()] = (qdf["_END_TS"] - qdf["_START_TS"]).dt.total_seconds() / 3600.0
        seg_cols = []
        for a, b in SEGMENTS:
            col = _segment_col_name(a, b)
            qdf[col] = (qdf[f"_RES_{b}"] - qdf[f"_RES_{a}"]).dt.total_seconds() / 3600.0
            seg_cols.append(col)
        agg_fn = "min" if shipment_agg.lower().startswith("ear") else "max"
        ship = qdf.groupby(group_cols, dropna=False)[[_journey_col_name()] + seg_cols].agg(agg_fn).reset_index()
        ship["LANE"] = ship.apply(lambda r: _make_lane(r["POL"], r["POD"]), axis=1)
        # carry through SHIPMENT_MONTH_YEAR for trend bucketing (NEW: a derived attribute we keep per shipment)
        month_map = qdf.groupby("MASTER_SHIPMENT_ID")["_RES_" + ORDERED_MILESTONES[0]].first().dt.strftime("%b %Y")
        ship["SHIPMENT_MONTH_YEAR"] = ship["MASTER_SHIPMENT_ID"].map(month_map).fillna("")
        return ship
    df["_QUALIFIED"] = pd.notna(df["_START_TS"]) & pd.notna(df["_END_TS"]) & (df["_END_TS"] >= df["_START_TS"])
    qdf = df[df["_QUALIFIED"]].copy()
    if qdf.empty:
        return pd.DataFrame(columns=group_cols + ["LANE", _journey_col_name()])
    qdf[_journey_col_name()] = (qdf["_END_TS"] - qdf["_START_TS"]).dt.total_seconds() / 3600.0
    agg_fn = "min" if shipment_agg.lower().startswith("ear") else "max"
    ship = qdf.groupby(group_cols, dropna=False)[[_journey_col_name()]].agg(agg_fn).reset_index()
    ship["LANE"] = ship.apply(lambda r: _make_lane(r["POL"], r["POD"]), axis=1)
    month_map = qdf.groupby("MASTER_SHIPMENT_ID")["_RES_" + ORDERED_MILESTONES[0]].first().dt.strftime("%b %Y")
    ship["SHIPMENT_MONTH_YEAR"] = ship["MASTER_SHIPMENT_ID"].map(month_map).fillna("")
    return ship


def compute_lane_and_carrier_counts(shipment_lt):
    if shipment_lt.empty:
        return (pd.DataFrame(columns=["Tenant Name", "Lane", "Shipments"]),
                pd.DataFrame(columns=["Tenant Name", "Carrier Name", "Carrier SCAC", "Shipments"]))
    lane_counts = (shipment_lt.groupby(["TENANT_NAME", "LANE"], dropna=False)["MASTER_SHIPMENT_ID"]
                   .nunique().reset_index()
                   .rename(columns={"TENANT_NAME": "Tenant Name", "LANE": "Lane", "MASTER_SHIPMENT_ID": "Shipments"})
                   .sort_values(["Shipments", "Lane"], ascending=[False, True]))
    carrier_counts = (shipment_lt.groupby(["TENANT_NAME", "CARRIER_NAME", "CARRIER_SCAC"], dropna=False)["MASTER_SHIPMENT_ID"]
                      .nunique().reset_index()
                      .rename(columns={"TENANT_NAME": "Tenant Name", "CARRIER_NAME": "Carrier Name",
                                       "CARRIER_SCAC": "Carrier SCAC", "MASTER_SHIPMENT_ID": "Shipments"})
                      .sort_values(["Shipments", "Carrier Name"], ascending=[False, True]))
    return lane_counts, carrier_counts


def apply_top_n_lanes_filter(shipment_lt, top_n):
    if shipment_lt.empty or top_n <= 0:
        return shipment_lt
    lane_vol = (shipment_lt.groupby(["TENANT_NAME", "LANE"], dropna=False)["MASTER_SHIPMENT_ID"]
                .nunique().reset_index().rename(columns={"MASTER_SHIPMENT_ID": "SHIPMENTS"}))
    top_lanes = (lane_vol.sort_values(["TENANT_NAME", "SHIPMENTS", "LANE"], ascending=[True, False, True])
                 .groupby("TENANT_NAME", dropna=False).head(top_n)[["TENANT_NAME", "LANE"]].drop_duplicates())
    return shipment_lt.merge(top_lanes, on=["TENANT_NAME", "LANE"], how="inner")


def _stats_for_series(series, percentile_p, include_percentile, min_vol_for_pct, prefix):
    s = series.dropna()
    vol = int(s.shape[0])
    keys = ["TOTAL_H", "TOTAL_D", "MIN_H", "MIN_D", "MED_H", "MED_D", "PCT_H", "PCT_D", "MAX_H", "MAX_D"]
    out = {f"{prefix}_{k}": None for k in keys}
    if vol == 0:
        return out
    total_h = _round_hours(float(s.sum()))
    min_h = _round_hours(float(s.min()))
    med_h = _round_hours(float(s.median()))
    max_h = _round_hours(float(s.max()))
    out[f"{prefix}_TOTAL_H"] = total_h; out[f"{prefix}_TOTAL_D"] = _round_days_from_hours(total_h)
    out[f"{prefix}_MIN_H"] = min_h; out[f"{prefix}_MIN_D"] = _round_days_from_hours(min_h)
    out[f"{prefix}_MED_H"] = med_h; out[f"{prefix}_MED_D"] = _round_days_from_hours(med_h)
    out[f"{prefix}_MAX_H"] = max_h; out[f"{prefix}_MAX_D"] = _round_days_from_hours(max_h)
    if include_percentile and vol >= int(min_vol_for_pct):
        pct_h = _round_hours(_safe_quantile(s, percentile_p / 100.0))
        out[f"{prefix}_PCT_H"] = pct_h; out[f"{prefix}_PCT_D"] = _round_days_from_hours(pct_h)
    return out


def build_duration_configs(start_ms, end_ms, whole_journey):
    configs = [{"data_col": _journey_col_name(), "prefix": "JOURNEY", "label": f"{start_ms}-{end_ms}", "display_mode": "journey"}]
    if whole_journey:
        for a, b in SEGMENTS:
            configs.append({"data_col": _segment_col_name(a, b), "prefix": f"SEG_{a}_{b}", "label": f"{a}-{b}", "display_mode": "segment"})
    return configs


def _group_stats(g, duration_configs, percentile_p, include_pct, min_vol):
    result = {"VOLUME": int(g["MASTER_SHIPMENT_ID"].nunique())}
    for cfg in duration_configs:
        result.update(_stats_for_series(g[cfg["data_col"]], percentile_p, include_pct, min_vol, cfg["prefix"]))
    return pd.Series(result)


def build_carrier_lane_report(shipment_lt, percentile_p, include_pct, min_vol_for_pct, duration_configs):
    base_cols = ["TENANT_NAME", "LANE", "CARRIER_NAME", "CARRIER_SCAC", "VOLUME", "_IS_LANE_ROW", "_POL", "_POD"]
    metric_cols = []
    for cfg in duration_configs:
        pfx = cfg["prefix"]
        metric_cols.extend([f"{pfx}_{k}" for k in ["TOTAL_H", "TOTAL_D", "MIN_H", "MIN_D", "MED_H", "MED_D", "PCT_H", "PCT_D", "MAX_H", "MAX_D"]])
    cols = base_cols + metric_cols
    if shipment_lt.empty:
        return pd.DataFrame(columns=cols)
    lane_stats = (shipment_lt.groupby(["TENANT_NAME", "POL", "POD", "LANE"], dropna=False)
                  .apply(lambda g: _group_stats(g, duration_configs, percentile_p, include_pct, min_vol_for_pct)).reset_index())
    lane_stats["CARRIER_NAME"] = "ALL CARRIERS"; lane_stats["CARRIER_SCAC"] = ""
    carrier_stats = (shipment_lt.groupby(["TENANT_NAME", "POL", "POD", "LANE", "CARRIER_NAME", "CARRIER_SCAC"], dropna=False)
                     .apply(lambda g: _group_stats(g, duration_configs, percentile_p, include_pct, min_vol_for_pct)).reset_index())
    lane_stats = lane_stats.sort_values(["TENANT_NAME", "VOLUME", "LANE"], ascending=[True, False, True])
    rows = []
    for _, lr in lane_stats.iterrows():
        tenant, lane, pol, pod = lr["TENANT_NAME"], lr["LANE"], lr["POL"], lr["POD"]
        lane_row = {"TENANT_NAME": tenant, "LANE": lane, "CARRIER_NAME": lr["CARRIER_NAME"],
                    "CARRIER_SCAC": lr["CARRIER_SCAC"], "VOLUME": lr["VOLUME"],
                    "_IS_LANE_ROW": True, "_POL": pol, "_POD": pod}
        for mc in metric_cols: lane_row[mc] = lr.get(mc)
        rows.append(lane_row)
        csub = carrier_stats[(carrier_stats["TENANT_NAME"] == tenant)
                             & (carrier_stats["POL"].astype(str) == str(pol))
                             & (carrier_stats["POD"].astype(str) == str(pod))].sort_values(["VOLUME", "CARRIER_NAME"], ascending=[False, True])
        for _, cr in csub.iterrows():
            crow = {"TENANT_NAME": tenant, "LANE": "", "CARRIER_NAME": cr["CARRIER_NAME"],
                    "CARRIER_SCAC": cr["CARRIER_SCAC"], "VOLUME": cr["VOLUME"],
                    "_IS_LANE_ROW": False, "_POL": pol, "_POD": pod}
            for mc in metric_cols: crow[mc] = cr.get(mc)
            rows.append(crow)
    return pd.DataFrame(rows, columns=cols)


def compute_insights_for_metric(shipment_lt, metric_cfg, percentile_p, pct_threshold_on, pct_threshold_pct,
                                rec_threshold_on, rec_threshold_pct):
    metric_col = metric_cfg["data_col"]
    if shipment_lt.empty or metric_col not in shipment_lt.columns:
        return pd.DataFrame(), pd.DataFrame()
    valid = shipment_lt.dropna(subset=[metric_col]).copy()
    if valid.empty:
        return pd.DataFrame(), pd.DataFrame()
    lane_base = (valid.groupby(["TENANT_NAME", "LANE"], dropna=False)
                 .agg(LANE_SHIPMENTS=("MASTER_SHIPMENT_ID", "nunique"),
                      LANE_MEDIAN_H=(metric_col, "median"),
                      CARRIER_COUNT=("CARRIER_NAME", "nunique")).reset_index())
    lane_base["PERCENTILE_MIN_SHIPMENTS_REQUIRED"] = lane_base["LANE_SHIPMENTS"].apply(
        lambda x: _pct_to_count(int(x), float(pct_threshold_pct)) if pct_threshold_on else 0)
    lane_base["RECOMMENDATION_MIN_SHIPMENTS_REQUIRED"] = lane_base["LANE_SHIPMENTS"].apply(
        lambda x: _pct_to_count(int(x), float(rec_threshold_pct)) if rec_threshold_on else 0)
    pxx_rows = []
    for _, row in lane_base.iterrows():
        sub = valid[(valid["TENANT_NAME"] == row["TENANT_NAME"]) & (valid["LANE"] == row["LANE"])][metric_col]
        if pct_threshold_on and int(row["LANE_SHIPMENTS"]) < int(row["PERCENTILE_MIN_SHIPMENTS_REQUIRED"]):
            pxx_rows.append(None)
        else:
            pxx_rows.append(_safe_quantile(sub, percentile_p / 100.0))
    lane_base["LANE_PXX_H"] = pxx_rows
    lane_base["LANE_MEDIAN_H"] = lane_base["LANE_MEDIAN_H"].apply(_round_hours)
    lane_base["LANE_MEDIAN_D"] = lane_base["LANE_MEDIAN_H"].apply(_round_days_from_hours)
    lane_base["LANE_PXX_H"] = lane_base["LANE_PXX_H"].apply(_round_hours)
    lane_base["LANE_PXX_D"] = lane_base["LANE_PXX_H"].apply(_round_days_from_hours)
    merged = valid.merge(lane_base[["TENANT_NAME", "LANE", "LANE_MEDIAN_H", "LANE_SHIPMENTS",
                                    "PERCENTILE_MIN_SHIPMENTS_REQUIRED", "RECOMMENDATION_MIN_SHIPMENTS_REQUIRED"]],
                         on=["TENANT_NAME", "LANE"], how="left")
    merged["ABS_DEV_H"] = (merged[metric_col] - merged["LANE_MEDIAN_H"]).abs()
    carrier_rows = []
    for (tenant, lane, cname, scac), g in merged.groupby(["TENANT_NAME", "LANE", "CARRIER_NAME", "CARRIER_SCAC"], dropna=False):
        shipments = int(g["MASTER_SHIPMENT_ID"].nunique())
        lane_ships = int(g["LANE_SHIPMENTS"].iloc[0]) if not g.empty else 0
        share = round((shipments / lane_ships * 100.0), 2) if lane_ships > 0 else None
        pct_min = int(g["PERCENTILE_MIN_SHIPMENTS_REQUIRED"].iloc[0]) if not g.empty else 0
        rec_min = int(g["RECOMMENDATION_MIN_SHIPMENTS_REQUIRED"].iloc[0]) if not g.empty else 0
        cs = g[metric_col].dropna(); ds = g["ABS_DEV_H"].dropna()
        c_med = _round_hours(float(cs.median())) if not cs.empty else None
        mad = _round_hours(float(ds.median())) if not ds.empty else None
        pct_elig = not pct_threshold_on or shipments >= pct_min
        rec_elig = not rec_threshold_on or shipments >= rec_min
        c_pxx = _round_hours(_safe_quantile(cs, percentile_p / 100.0)) if pct_elig and not cs.empty else None
        d_pxx = _round_hours(_safe_quantile(ds, percentile_p / 100.0)) if pct_elig and not ds.empty else None
        carrier_rows.append({"TENANT_NAME": tenant, "LANE": lane, "CARRIER_NAME": cname, "CARRIER_SCAC": scac,
                             "SHIPMENTS": shipments, "CARRIER_SHARE_PCT": share,
                             "CARRIER_MEDIAN_H": c_med, "CARRIER_MEDIAN_D": _round_days_from_hours(c_med),
                             "CARRIER_PXX_H": c_pxx, "CARRIER_PXX_D": _round_days_from_hours(c_pxx),
                             "MEDIAN_ABS_DEV_H": mad, "MEDIAN_ABS_DEV_D": _round_days_from_hours(mad),
                             "PXX_ABS_DEV_H": d_pxx, "PXX_ABS_DEV_D": _round_days_from_hours(d_pxx),
                             "RECOMMENDATION_ELIGIBLE": rec_elig, "PERCENTILE_ELIGIBLE": pct_elig})
    recs = pd.DataFrame(carrier_rows)
    if recs.empty:
        recs["RANK_IN_LANE"] = pd.Series(dtype="Int64")
        return lane_base, recs
    recs["MAD_SORT"] = recs["MEDIAN_ABS_DEV_H"].fillna(np.inf)
    recs["PXX_SORT"] = recs["PXX_ABS_DEV_H"].fillna(np.inf)
    elig = recs[recs["RECOMMENDATION_ELIGIBLE"]].copy()
    elig = elig.sort_values(["TENANT_NAME", "LANE", "MAD_SORT", "PXX_SORT", "SHIPMENTS", "CARRIER_NAME"],
                            ascending=[True, True, True, True, False, True])
    elig["RANK_IN_LANE"] = elig.groupby(["TENANT_NAME", "LANE"], dropna=False).cumcount() + 1
    non_elig = recs[~recs["RECOMMENDATION_ELIGIBLE"]].copy()
    non_elig["RANK_IN_LANE"] = pd.NA
    out = pd.concat([elig, non_elig], ignore_index=True).drop(columns=["MAD_SORT", "PXX_SORT"], errors="ignore")
    lane_summary = lane_base.sort_values(["LANE_SHIPMENTS", "LANE"], ascending=[False, True])
    return lane_summary, out


# =====================================================================
# NEW DERIVED HELPERS — operate on existing outputs (no core logic change)
# =====================================================================

def compute_reliability_score(lane_median: float, lane_mad: float,
                              all_medians: List[float], all_mads: List[float],
                              w_consistency: float = 0.6, w_speed: float = 0.4) -> Dict:
    """Composite 0-100 score per lane: w1*consistency + w2*speed. Both normalized min-max across network."""
    def norm(x, vals):
        vals = [v for v in vals if v is not None and not (isinstance(v, float) and np.isnan(v))]
        if not vals or x is None:
            return 0.5
        lo, hi = min(vals), max(vals)
        if hi == lo:
            return 0.5
        return (x - lo) / (hi - lo)
    consistency = 100.0 * (1.0 - norm(lane_mad, all_mads))
    speed = 100.0 * (1.0 - norm(lane_median, all_medians))
    score = w_consistency * consistency + w_speed * speed
    return {"score": round(score), "consistency": round(consistency), "speed": round(speed)}


def compute_skew_analysis(lane_shipments_series: pd.Series, lane_median: float) -> List[Dict]:
    """For each trim % in {5, 10, 20}: trim slowest N%, recompute median."""
    s = lane_shipments_series.dropna().sort_values(ascending=False)
    n = len(s)
    out = []
    for pct in [5, 10, 20]:
        n_trim = int(np.ceil(n * pct / 100.0))
        trimmed = s.iloc[n_trim:]
        new_med_h = float(trimmed.median()) if not trimmed.empty else lane_median
        new_med_d = round(new_med_h / 24.0, 1)
        out.append({"pct": pct, "count": n_trim, "new_median_d": new_med_d})
    return out


def compute_trend_buckets(shipment_lt: pd.DataFrame, lane: str, granularity: str) -> List[Dict]:
    """Group shipments by SHIPMENT_MONTH_YEAR (Month) or Quarter, compute median per bucket."""
    if shipment_lt.empty or "SHIPMENT_MONTH_YEAR" not in shipment_lt.columns:
        return []
    sub = shipment_lt[shipment_lt["LANE"] == lane].copy()
    if sub.empty or sub["SHIPMENT_MONTH_YEAR"].eq("").all():
        return []
    if granularity == "Month":
        sub["BUCKET"] = sub["SHIPMENT_MONTH_YEAR"]
    else:
        # Quarter derivation from "Jan 2026" → "Q1 2026"
        def to_quarter(s):
            if not s: return ""
            try:
                m_str, y_str = s.split()
                month_to_q = {"Jan": "Q1", "Feb": "Q1", "Mar": "Q1", "Apr": "Q2", "May": "Q2", "Jun": "Q2",
                              "Jul": "Q3", "Aug": "Q3", "Sep": "Q3", "Oct": "Q4", "Nov": "Q4", "Dec": "Q4"}
                return f"{month_to_q.get(m_str, '')} {y_str}"
            except Exception:
                return ""
        sub["BUCKET"] = sub["SHIPMENT_MONTH_YEAR"].apply(to_quarter)
    sub = sub[sub["BUCKET"] != ""]
    if sub.empty: return []
    grouped = sub.groupby("BUCKET")[_journey_col_name()].median().reset_index()
    grouped["median_d"] = (grouped[_journey_col_name()] / 24.0).round(1)
    # Sort buckets chronologically
    def sort_key(b):
        try:
            if b.startswith("Q"):
                q, y = b.split(); return int(y) * 10 + int(q[1])
            m, y = b.split()
            month_idx = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,"Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12}
            return int(y) * 100 + month_idx.get(m, 0)
        except Exception:
            return 0
    grouped["_sort"] = grouped["BUCKET"].apply(sort_key)
    grouped = grouped.sort_values("_sort")
    return [{"bucket": r["BUCKET"], "median_d": float(r["median_d"])} for _, r in grouped.iterrows()]


def compute_bottleneck(segment_stats: List[Dict]) -> Optional[Dict]:
    """Find segment with highest P-XX/median ratio."""
    candidates = [s for s in segment_stats if s.get("median_d") and s.get("pxx_d") and s["median_d"] > 0]
    if not candidates:
        return None
    for s in candidates:
        s["ratio"] = s["pxx_d"] / s["median_d"]
    worst = max(candidates, key=lambda x: x["ratio"])
    worst["improvement_d"] = round(worst["pxx_d"] - worst["median_d"], 1)
    return worst


def classify_concentration(top_share: float, top_two_share: float,
                           threshold_high: float = 60.0, threshold_moderate: float = 40.0) -> Dict:
    if top_share is None:
        return {"level": "Unknown", "color": "gray"}
    if top_share > threshold_high:
        return {"level": "High", "color": "high", "label": f"High concentration risk · top carrier handles {top_share}%"}
    if top_share >= threshold_moderate:
        return {"level": "Moderate", "color": "moderate", "label": f"Moderate concentration · top 2 handle {top_two_share}%"}
    return {"level": "Low", "color": "low", "label": f"Low concentration · top 2 handle {top_two_share}%"}


def compute_problematic_shipments(shipment_lt: pd.DataFrame, lane: str, scope_col: str,
                                  lane_median_h: float, lane_pxx_h: float, mode: str,
                                  limit: int = 50) -> pd.DataFrame:
    """Return shipments in lane where lead exceeds either lane median or lane PXX, sorted by largest exceedance.
       scope_col is JOURNEY_LEAD_HOURS for whole-journey, or SEG_*_HOURS for a specific segment."""
    if shipment_lt.empty or scope_col not in shipment_lt.columns:
        return pd.DataFrame()
    sub = shipment_lt[shipment_lt["LANE"] == lane].copy()
    if sub.empty:
        return sub
    sub = sub.dropna(subset=[scope_col])
    threshold = lane_median_h if mode == "Above Median" else lane_pxx_h
    if threshold is None:
        return pd.DataFrame()
    problematic = sub[sub[scope_col] > threshold].copy()
    if problematic.empty:
        return problematic
    problematic["DELTA_H"] = problematic[scope_col] - threshold
    problematic["LEAD_D"] = (problematic[scope_col] / 24.0).round(1)
    problematic["DELTA_D"] = (problematic["DELTA_H"] / 24.0).round(1)
    problematic = problematic.sort_values("DELTA_H", ascending=False).head(limit)
    return problematic[["MASTER_SHIPMENT_ID", "CARRIER_NAME", "CARRIER_SCAC", "LEAD_D", "DELTA_D"]]


# =====================================================================
# EXCEL WRITERS — verbatim from your app.py
# =====================================================================
def build_export_rename_map(duration_configs, percentile_p):
    export_cols = {"TENANT_NAME": DISPLAY_COLS["TENANT_NAME"], "LANE": DISPLAY_COLS["LANE"],
                   "CARRIER_NAME": DISPLAY_COLS["CARRIER_NAME"], "CARRIER_SCAC": DISPLAY_COLS["CARRIER_SCAC"],
                   "VOLUME": DISPLAY_COLS["VOLUME"]}
    for cfg in duration_configs:
        pfx = cfg["prefix"]; label = cfg["label"]
        if cfg["display_mode"] == "journey":
            for k in ["TOTAL_H", "TOTAL_D", "MIN_H", "MIN_D", "MED_H", "MED_D", "MAX_H", "MAX_D"]:
                export_cols[f"{pfx}_{k}"] = DISPLAY_COLS[k]
            export_cols[f"{pfx}_PCT_H"] = DISPLAY_COLS["PCT_H"].format(p=percentile_p)
            export_cols[f"{pfx}_PCT_D"] = DISPLAY_COLS["PCT_D"].format(p=percentile_p)
        else:
            for k, friendly in [("TOTAL_H", "Total Lead Time (Hours)"), ("TOTAL_D", "Total Lead Time (Days)"),
                                ("MIN_H", "Min Lead Time (Hours)"), ("MIN_D", "Min Lead Time (Days)"),
                                ("MED_H", "Median Lead Time (Hours)"), ("MED_D", "Median Lead Time (Days)"),
                                ("MAX_H", "Max Lead Time (Hours)"), ("MAX_D", "Max Lead Time (Days)")]:
                export_cols[f"{pfx}_{k}"] = f"{label} {friendly}"
            export_cols[f"{pfx}_PCT_H"] = f"{label} P{percentile_p} Lead Time (Hours)"
            export_cols[f"{pfx}_PCT_D"] = f"{label} P{percentile_p} Lead Time (Days)"
    return export_cols


def write_excel_counts(lane_counts, carrier_counts):
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        lane_counts.to_excel(writer, sheet_name="Lane Counts", index=False)
        carrier_counts.to_excel(writer, sheet_name="Carrier Counts", index=False)
        for sheet in ["Lane Counts", "Carrier Counts"]:
            ws = writer.book[sheet]
            for cell in ws[1]: cell.font = Font(bold=True)
            for col in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col)].width = 26
    output.seek(0); return output.getvalue()


def write_excel_final(raw_df, report_df, duration_configs, percentile_p):
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    output = io.BytesIO()
    raw_export = add_shipment_month_year(raw_df).drop(columns=["_FIRST_TS"], errors="ignore")
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        raw_export.to_excel(writer, sheet_name="Raw Data", index=False)
        export_map = build_export_rename_map(duration_configs, percentile_p)
        if report_df.empty:
            pd.DataFrame(columns=list(export_map.values())).to_excel(writer, sheet_name="Carrier Lane Lead", index=False)
        else:
            df = report_df.copy()
            lane_flags = df["_IS_LANE_ROW"].astype(bool).to_list()
            df = df.drop(columns=["_IS_LANE_ROW", "_POL", "_POD"], errors="ignore")
            keys = [k for k in export_map.keys() if k in df.columns]
            df = df[keys].rename(columns=export_map)
            df.to_excel(writer, sheet_name="Carrier Lane Lead", index=False)
            ws = writer.book["Carrier Lane Lead"]
            bold = Font(bold=True)
            for cell in ws[1]: cell.font = bold
            lane_idx = list(df.columns).index(DISPLAY_COLS["LANE"]) + 1
            for i, is_lane in enumerate(lane_flags, start=2):
                if is_lane: ws.cell(row=i, column=lane_idx).font = bold
            for idx in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(idx)].width = 24
        for cell in writer.book["Raw Data"][1]: cell.font = Font(bold=True)
    output.seek(0); return output.getvalue()


def write_problematic_excel(problematic_df: pd.DataFrame, scope_label: str, mode: str) -> bytes:
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        out_df = problematic_df.rename(columns={
            "MASTER_SHIPMENT_ID": "Shipment ID", "CARRIER_NAME": "Carrier",
            "CARRIER_SCAC": "SCAC", "LEAD_D": "Lead Time (Days)",
            "DELTA_D": f"Days Over ({mode})"})
        out_df.to_excel(writer, sheet_name="Problematic", index=False)
        ws = writer.book["Problematic"]
        for cell in ws[1]: cell.font = Font(bold=True)
        for idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 24
        # Add a header row above with the scope context
    output.seek(0); return output.getvalue()


# =====================================================================
# UI HELPERS (HTML rendering)
# =====================================================================
def render_journey_diagram(start_ms: str, end_ms: str, whole_journey: bool):
    """Render the 8-milestone flow with the active path highlighted."""
    start_idx = ORDERED_MILESTONES.index(start_ms)
    end_idx = ORDERED_MILESTONES.index(end_ms)
    track_left = 6 + (start_idx / 7) * 88 if not whole_journey else 6
    track_right = 6 + ((7 - end_idx) / 7) * 88 if not whole_journey else 6
    milestones_html = ""
    for i, ms in enumerate(ORDERED_MILESTONES):
        full, place = MILESTONE_LONG[ms]
        if whole_journey:
            cls = "endpoint" if i in (0, 7) else "in-path"
        else:
            if i == start_idx or i == end_idx:
                cls = "endpoint"
            elif start_idx < i < end_idx:
                cls = "in-path"
            else:
                cls = ""
        dot_label = str(i + 1)
        place_cls = "pol" if place == "POL" else "pod"
        milestones_html += f"""
        <div class="v7-jd-milestone {cls}">
          <div class="dot">{dot_label}</div>
          <div class="code">{ms}</div>
          <div class="label">{full.replace(' ', '<br>', 1)}</div>
          <div class="place {place_cls}">{place}</div>
        </div>"""
    return f"""
    <div class="v7-journey-diagram">
      <div class="v7-jd-port-headers">
        <div class="pol">◀ Port of Loading (POL)</div>
        <div class="transit">Ocean Transit</div>
        <div class="pod">Port of Discharge (POD) ▶</div>
      </div>
      <div class="v7-jd-flow">
        <div class="v7-jd-track-bg"></div>
        <div class="v7-jd-track-active" style="left: {track_left}%; right: {track_right}%;"></div>
        {milestones_html}
      </div>
    </div>
    """


def render_whole_journey_card(active: bool, median_d, pxx_d, mad_d, n_ships: int):
    cls = "" if active else "inactive"
    sel_pill = '<div class="wj-selected-pill">● Selected</div>' if active else ""
    def fmt(v, suffix="d"):
        return f"{v}<span class='u'>{suffix}</span>" if v is not None else "—"
    return f"""
    <div class="v7-jd-whole-journey {cls}">
      <div class="wj-icon">⛴</div>
      <div class="wj-label">
        <div class="l">End-to-end summary</div>
        <div class="v">Whole Journey · CEP → CER</div>
      </div>
      <div class="wj-stat"><div class="l">Median</div><div class="v">{fmt(median_d)}</div></div>
      <div class="wj-stat"><div class="l">P80</div><div class="v">{fmt(pxx_d)}</div></div>
      <div class="wj-stat"><div class="l">MAD</div><div class="v">{fmt(mad_d)}</div></div>
      <div class="wj-stat"><div class="l">Shipments</div><div class="v">{n_ships:,}</div></div>
      {sel_pill}
    </div>
    """


def render_kpi_card(label: str, value: str, sub: str, mini_pct: Optional[float] = None, tooltip: str = ""):
    mini_html = (f'<div class="v7-mini-bar"><div class="v7-mini-bar-fill" style="width:{mini_pct}%"></div></div>'
                 if mini_pct is not None else "")
    return f"""
    <div class="v7-kpi-card" title="{tooltip}">
      <div class="label">{label}</div>
      <div class="value">{value}</div>
      {mini_html}
      <div class="sub">{sub}</div>
    </div>
    """


def render_best_carrier(carrier_row, lane_median_d):
    if carrier_row is None:
        return '<div class="v7-insight-card"><div class="v7-ic-head"><div class="v7-ic-title">Best Carrier</div></div><p style="color: var(--text-tertiary);">No eligible carrier in this lane.</p></div>'
    name = carrier_row.get("CARRIER_NAME", "—")
    scac = carrier_row.get("CARRIER_SCAC", "")
    c_med = carrier_row.get("CARRIER_MEDIAN_D")
    share = carrier_row.get("CARRIER_SHARE_PCT")
    mad = carrier_row.get("MEDIAN_ABS_DEV_D")
    c_pxx = carrier_row.get("CARRIER_PXX_D")
    delta = (lane_median_d - c_med) if (lane_median_d is not None and c_med is not None) else None
    delta_str = f"{abs(delta):.1f}d {'faster' if delta and delta > 0 else 'slower'}" if delta is not None else "—"
    return f"""
    <div class="v7-insight-card featured">
      <div class="v7-ic-head"><div class="v7-ic-title">Best Carrier <span class="v7-tag-ai">AI Pick</span></div></div>
      <div class="v7-bc-rank">🏆 Rank #1</div>
      <div class="v7-bc-name">{name}</div>
      <div class="v7-bc-scac">SCAC: {scac}</div>
      <div class="v7-bc-quote"><strong>{name}</strong> delivers <strong>{delta_str}</strong> than the lane median, with MAD of <strong>{mad if mad is not None else '—'}d</strong>.</div>
      <div class="v7-bc-stats">
        <div class="v7-bc-stat"><div class="l">Median</div><div class="v">{c_med if c_med is not None else '—'}d</div></div>
        <div class="v7-bc-stat"><div class="l">Share</div><div class="v">{share if share is not None else '—'}%</div></div>
        <div class="v7-bc-stat"><div class="l">MAD</div><div class="v">{mad if mad is not None else '—'}d</div></div>
        <div class="v7-bc-stat"><div class="l">P-XX</div><div class="v">{c_pxx if c_pxx is not None else '—'}d</div></div>
      </div>
    </div>
    """


def render_top5(carrier_rows, lane_median_d):
    if carrier_rows.empty:
        return '<div class="v7-insight-card"><div class="v7-ic-head"><div class="v7-ic-title">Top 5 Recommended</div></div><p style="color: var(--text-tertiary);">No eligible carriers.</p></div>'
    rows = ""
    for _, r in carrier_rows.head(5).iterrows():
        rank = int(r["RANK_IN_LANE"]) if pd.notna(r.get("RANK_IN_LANE")) else "—"
        c_med = r.get("CARRIER_MEDIAN_D")
        mad = r.get("MEDIAN_ABS_DEV_D")
        share = r.get("CARRIER_SHARE_PCT")
        delta = (lane_median_d - c_med) if (lane_median_d is not None and c_med is not None) else None
        delta_str = f"{'−' if delta and delta > 0 else '+'}{abs(delta):.1f}d vs lane" if delta is not None else "—"
        cls = "rank-1" if rank == 1 else ""
        rows += f"""
        <div class="v7-rec-row {cls}">
          <div class="rank">{rank}</div>
          <div class="info"><div class="name">{r['CARRIER_NAME']} <span style="color: var(--text-quaternary); font-size:10px;">{r['CARRIER_SCAC']}</span></div><div class="why">MAD {mad if mad is not None else '—'}d · {delta_str}</div></div>
          <div class="stats">
            <div class="stat"><div class="v">{c_med if c_med is not None else '—'}d</div><div class="l">Median</div></div>
            <div class="stat"><div class="v">{share if share is not None else '—'}%</div><div class="l">Share</div></div>
          </div>
        </div>"""
    return f"""
    <div class="v7-insight-card">
      <div class="v7-ic-head"><div class="v7-ic-title">Top 5 Recommended</div></div>
      <div class="v7-rec-list">{rows}</div>
    </div>"""


def render_problematic_card(scope_label: str, prob_df: pd.DataFrame, mode: str,
                             lane_median_d, trimmed_d):
    if prob_df.empty:
        prob_html = '<p style="color: var(--text-tertiary); font-size: 12px;">No problematic shipments for this scope.</p>'
    else:
        rows = ""
        for i, (_, r) in enumerate(prob_df.head(8).iterrows()):
            cls = "" if i < 3 else "med"
            rows += f"""
            <div class="v7-prob-row {cls}">
              <div class="shipment-id"><span class="c">{r['CARRIER_NAME']}</span>{r['MASTER_SHIPMENT_ID']}</div>
              <div class="actual">{r['LEAD_D']}<span style="font-weight:400; font-size:11px; color: var(--text-tertiary);">d</span></div>
              <div class="delta">+{r['DELTA_D']}d</div>
            </div>"""
        prob_html = f'<div class="v7-prob-list">{rows}</div>'
    count = len(prob_df)
    footer = (f"<strong>{count} shipments</strong> {'above ' + mode.lower().replace('above ', '')} · "
              f"trimming top 5% brings median from {lane_median_d}d to <strong>{trimmed_d}d</strong>")
    return f"""
    <div class="v7-insight-card">
      <div class="v7-ic-head"><div class="v7-ic-title">Problematic Shipments</div></div>
      <div class="v7-prob-scope">
        <div class="l">Currently scoped to</div>
        <div class="v">{scope_label}</div>
      </div>
      {prob_html}
      <div class="v7-prob-footer">{footer}</div>
    </div>"""


def render_skew_card(skew_rows: List[Dict], lane_median_d):
    if not skew_rows:
        return '<div class="v7-insight-card"><div class="v7-ic-head"><div class="v7-ic-title">Median Skew Analysis</div></div><p style="color: var(--text-tertiary);">Not enough data.</p></div>'
    max_drop = max((lane_median_d - r["new_median_d"]) for r in skew_rows) if lane_median_d else 1
    rows_html = ""
    for r in skew_rows:
        drop = (lane_median_d - r["new_median_d"]) if lane_median_d else 0
        pct_bar = min(100, (drop / max_drop * 100)) if max_drop > 0 else 0
        rows_html += f"""
        <div class="v7-skew-row">
          <div class="what">Drop top {r['pct']}%<span class="count">{r['count']} ships</span></div>
          <div class="bar"><div class="fill" style="width: {pct_bar}%;"></div></div>
          <div class="result">{r['new_median_d']}d<span class="from">from {lane_median_d}d</span></div>
        </div>"""
    return f"""
    <div class="v7-insight-card">
      <div class="v7-ic-head"><div class="v7-ic-title">Median Skew Analysis</div></div>
      {rows_html}
    </div>"""


def render_reliability_card(score_info: Dict, n_lanes: int, rank: int):
    score = score_info["score"]; cons = score_info["consistency"]; spd = score_info["speed"]
    return f"""
    <div class="v7-insight-card">
      <div class="v7-ic-head"><div class="v7-ic-title">Lane Reliability Score <span class="v7-tag-ai">Derived</span></div></div>
      <div style="display: grid; grid-template-columns: 130px 1fr; gap: 18px; align-items: center; margin-bottom: 14px;">
        <div style="width: 130px; height: 130px; position: relative;">
          <svg width="130" height="130" viewBox="0 0 130 130" style="transform: rotate(-90deg);">
            <circle cx="65" cy="65" r="55" stroke="var(--border)" stroke-width="11" fill="none"/>
            <circle cx="65" cy="65" r="55" stroke="url(#scoreGrad)" stroke-width="11" fill="none" stroke-dasharray="345" stroke-dashoffset="{345 - (345 * score / 100)}" stroke-linecap="round"/>
            <defs><linearGradient id="scoreGrad" x1="0%" y1="0%" x2="100%" y2="100%"><stop offset="0%" stop-color="#6366f1"/><stop offset="100%" stop-color="#ec4899"/></linearGradient></defs>
          </svg>
          <div style="position: absolute; inset: 0; display: grid; place-items: center; font-size: 32px; font-weight: 800; letter-spacing: -0.02em; color: var(--text);">{score}<span style="font-size: 11px; color: var(--text-tertiary); font-weight: 400; margin-left: 2px;">/100</span></div>
        </div>
        <div>
          <div style="font-size: 12.5px; color: var(--text-secondary); margin-bottom: 10px;">This lane ranks <strong style="color: var(--accent-light);">#{rank} of your {n_lanes} lanes</strong> on reliability.</div>
          <div style="display: flex; justify-content: space-between; font-size: 10.5px; color: var(--text-tertiary); margin-bottom: 6px;"><span>Consistency (60% weight)</span><span>{cons}</span></div>
          <div style="height: 6px; background: var(--bg-elev); border-radius: 3px; overflow: hidden; margin-bottom: 4px;"><div style="height: 100%; background: var(--green); width: {cons}%;"></div></div>
          <div style="display: flex; justify-content: space-between; font-size: 10.5px; color: var(--text-tertiary); margin-bottom: 6px;"><span>Speed (40% weight)</span><span>{spd}</span></div>
          <div style="height: 6px; background: var(--bg-elev); border-radius: 3px; overflow: hidden;"><div style="height: 100%; background: var(--blue); width: {spd}%;"></div></div>
        </div>
      </div>
      <div class="v7-reliability-explanation">
        <div class="v7-re-head">How is this score calculated?</div>
        <div class="v7-re-body">
          <p>The score combines two simple things about this lane:</p>
          <ul>
            <li><strong>Consistency</strong> — how tightly your shipments cluster around the typical case (from <strong>MAD</strong>). Lower MAD = higher consistency.</li>
            <li><strong>Speed</strong> — how fast this lane is compared to all your other lanes (from the <strong>lane median</strong>). Lower median = higher speed.</li>
          </ul>
          <p>Both are scored 0–100 by comparing this lane against your network. Final score = <strong>60% Consistency + 40% Speed</strong>.</p>
        </div>
      </div>
    </div>"""


def render_analyst_table(report_df: pd.DataFrame, percentile_p: int, max_rows: int = 60):
    if report_df.empty:
        return '<div class="v7-analyst-table" style="padding: 20px; color: var(--text-tertiary);">No data to display.</div>'
    pxx_col = f"P{percentile_p}"
    rows_html = ""
    df = report_df.head(max_rows)
    for _, r in df.iterrows():
        is_lane = bool(r.get("_IS_LANE_ROW", False))
        cls = "lane-row" if is_lane else "carrier-row"
        first = r["LANE"] if is_lane else r["CARRIER_NAME"]
        scac = "—" if is_lane else r["CARRIER_SCAC"]
        vol = f"{int(r['VOLUME']):,}" if pd.notna(r.get("VOLUME")) else "—"
        med = r.get("JOURNEY_MED_D"); pxx = r.get("JOURNEY_PCT_D")
        mx = r.get("JOURNEY_MAX_D")
        rows_html += f"""<tr class="{cls}">
          <td>{first}</td><td>{scac}</td><td>{vol}</td>
          <td>{med if pd.notna(med) else '—'}d</td>
          <td>{pxx if pd.notna(pxx) else '—'}d</td>
          <td>{mx if pd.notna(mx) else '—'}d</td>
        </tr>"""
    return f"""
    <div class="v7-analyst-table">
      <table>
        <thead><tr><th>Lane / Carrier</th><th>SCAC</th><th>Volume</th><th>Median</th><th>{pxx_col}</th><th>Max</th></tr></thead>
        <tbody>{rows_html}</tbody>
      </table>
    </div>"""


# =====================================================================
# MAIN APP
# =====================================================================
st.set_page_config(page_title="Lead Time Optimization · v7", layout="wide", initial_sidebar_state="collapsed")

# ---------------------------------------------------------------------
# HTML render fix: Streamlit's st.markdown treats lines with 4+ leading
# spaces as code blocks even with unsafe_allow_html=True. We monkey-patch
# st.markdown to aggressively strip per-line leading whitespace from any
# HTML body. This is safe because our HTML/CSS doesn't depend on
# indentation, and the JS we ship is single-line.
# ---------------------------------------------------------------------
_orig_markdown = st.markdown
def _patched_markdown(body, *args, **kwargs):
    if kwargs.get("unsafe_allow_html") and isinstance(body, str):
        body = "\n".join(line.lstrip() for line in body.split("\n")).strip()
    return _orig_markdown(body, *args, **kwargs)
st.markdown = _patched_markdown

st.markdown(V7_CSS, unsafe_allow_html=True)

# Session state init
if "raw_df" not in st.session_state: st.session_state["raw_df"] = None
if "file_name" not in st.session_state: st.session_state["file_name"] = None
if "active_scope" not in st.session_state: st.session_state["active_scope"] = "whole"  # "whole" or e.g. "VDL-VAD"
if "prob_mode" not in st.session_state: st.session_state["prob_mode"] = "Above Median"

# =====================================================================
# PHASE 1: FILE UPLOAD LANDING
# =====================================================================
if st.session_state["raw_df"] is None:
    st.markdown("""
    <div class="v7-upload-hero">
      <div class="logo">⬢</div>
      <div class="tag">Lead Time Optimization · Ocean</div>
      <h1>Upload your ocean shipment file</h1>
      <p>CSV or Excel export from Snowflake with the required columns. Once you upload, the full analysis loads and you can drill into segments, carriers, and problematic shipments.</p>
    </div>
    """, unsafe_allow_html=True)
    uploaded = st.file_uploader(" ", type=["csv", "xlsx", "xls"], label_visibility="collapsed")
    if uploaded is not None:
        try:
            df = _read_input(uploaded)
            st.session_state["raw_df"] = df
            st.session_state["file_name"] = uploaded.name
            st.rerun()
        except Exception as e:
            st.error(str(e))
    st.markdown("""
    <div style="text-align: center; margin-top: 32px; font-size: 12px; color: var(--text-quaternary);">
      Required columns: TENANT_NAME, MASTER_SHIPMENT_ID, POL, POD, CARRIER_NAME, CARRIER_SCAC, plus the milestone timestamps you want to analyze (CEP, CGI, CLL, VDL, VAD, CDD, CGO, CER — or their _P44 fallbacks for VDL/VAD).
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# =====================================================================
# PHASE 2: V7 UI (after file uploaded)
# =====================================================================
raw_df = st.session_state["raw_df"]

# ---- Topbar ----
col_a, col_b = st.columns([3, 2])
with col_a:
    st.markdown(f"""
    <div class="v7-breadcrumb"><span>Analytics</span><span class="sep">/</span><span>Ocean</span><span class="sep">/</span><span class="current">Lead Time Optimization</span></div>
    <div class="v7-page-title">Lead Time Optimization</div>
    <div style="font-size: 12px; color: var(--text-tertiary); margin-top: 4px;">File: <code style="color: var(--accent-light); font-family: 'SF Mono', monospace;">{st.session_state['file_name']}</code> · {raw_df.shape[0]:,} rows × {raw_df.shape[1]:,} cols</div>
    """, unsafe_allow_html=True)
with col_b:
    c1, c2 = st.columns([2, 1])
    with c1:
        if st.button("🔄 Upload a different file", use_container_width=True):
            st.session_state["raw_df"] = None; st.rerun()
    with c2:
        st.button("Save View", type="primary", use_container_width=True)

st.markdown("<div style='height: 12px'></div>", unsafe_allow_html=True)

# ---- Configure Strip ----
st.markdown('<div class="v7-configure-strip">', unsafe_allow_html=True)
cc1, cc2, cc3, cc4 = st.columns([2, 1, 1.4, 1])
with cc1:
    sc1, sc2 = st.columns(2)
    with sc1:
        start_ms = st.selectbox("Journey start", MILESTONES, index=MILESTONES.index("VDL"))
    with sc2:
        end_ms = st.selectbox("Journey end", MILESTONES, index=MILESTONES.index("VAD"))
with cc2:
    percentile_p = st.number_input("Percentile P-XX", min_value=1, max_value=99, value=80, step=1)
with cc3:
    lane_filter_option = st.selectbox("Lane filter", ["All lanes", "Top 10 by volume", "Top 25 by volume", "Top 50 by volume"])
    top_n_lanes = {"All lanes": 0, "Top 10 by volume": 10, "Top 25 by volume": 25, "Top 50 by volume": 50}[lane_filter_option]
with cc4:
    st.markdown("<div style='height: 8px'></div>", unsafe_allow_html=True)
    st.markdown(f"<div style='font-size: 11px; color: var(--text-tertiary); padding: 6px 0;'>Lane filter applies to all calculations</div>", unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)

# ---- Validation ----
missing = get_missing_columns(raw_df, start_ms, end_ms, whole_journey=True)
if missing:
    st.error("Required columns missing:\n\n- " + "\n- ".join(missing))
    st.stop()

# ---- Compute shipment-level data (for both modes — easier for the UI) ----
@st.cache_data(show_spinner=False)
def _compute_shipment_lt(raw_df_hash, start_ms, end_ms, agg, whole, top_n):
    raw = st.session_state["raw_df"]
    ship = compute_shipment_leadtimes(raw, start_ms, end_ms, agg, whole)
    ship = apply_top_n_lanes_filter(ship, top_n)
    return ship

raw_hash = hash((id(raw_df), raw_df.shape[0], raw_df.shape[1]))
shipment_lt = _compute_shipment_lt(raw_hash, start_ms, end_ms, "Earliest", True, top_n_lanes)

total_in_file = raw_df["MASTER_SHIPMENT_ID"].nunique() if "MASTER_SHIPMENT_ID" in raw_df.columns else 0
eligible = shipment_lt["MASTER_SHIPMENT_ID"].nunique() if not shipment_lt.empty else 0
coverage = (eligible / total_in_file * 100.0) if total_in_file else 0
lane_counts, carrier_counts = compute_lane_and_carrier_counts(shipment_lt)

# ---- KPI Strip ----
k1, k2, k3, k4 = st.columns(4)
with k1:
    st.markdown(render_kpi_card("Total Shipments", f"{total_in_file:,}", "unique master shipments in file",
                                tooltip="Unique MASTER_SHIPMENT_ID count"), unsafe_allow_html=True)
    raw_with_month = add_shipment_month_year(raw_df).drop(columns=["_FIRST_TS"], errors="ignore")
    raw_excel = io.BytesIO()
    with pd.ExcelWriter(raw_excel, engine="openpyxl") as w: raw_with_month.to_excel(w, sheet_name="Raw Data", index=False)
    raw_excel.seek(0)
    st.download_button("↓ Download Raw", raw_excel.getvalue(), file_name="raw_shipments.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_raw", use_container_width=True)
with k2:
    st.markdown(render_kpi_card("Eligible Shipments", f"{eligible:,}", f"{coverage:.1f}% coverage",
                                mini_pct=coverage), unsafe_allow_html=True)
with k3:
    n_lanes = lane_counts.shape[0]
    st.markdown(render_kpi_card("Unique Lanes", f"{n_lanes:,}",
                                f"top {min(6, n_lanes)} carry the majority"), unsafe_allow_html=True)
    counts_excel = write_excel_counts(lane_counts, carrier_counts)
    st.download_button("↓ Download Lane Counts", counts_excel, file_name="lane_carrier_counts.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_lanes", use_container_width=True)
with k4:
    n_carriers = carrier_counts.shape[0]
    st.markdown(render_kpi_card("Unique Carriers", f"{n_carriers:,}", "all carriers in eligible set"), unsafe_allow_html=True)
    st.download_button("↓ Download Carrier Counts", counts_excel, file_name="lane_carrier_counts.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_carriers", use_container_width=True)

st.markdown("<div style='height: 18px'></div>", unsafe_allow_html=True)

# ---- Journey Card ----
st.markdown('<div class="v7-journey-card">', unsafe_allow_html=True)
jh_left, jh_right = st.columns([3, 2])
with jh_left:
    st.markdown(f"""
    <div class="v7-jh-eyebrow">The Journey</div>
    <div class="v7-jh-title">Where is your lead time going?</div>
    <div class="v7-jh-subtitle">Whole Journey shows all 7 segments + an end-to-end summary card. Pick a segment to scope insights to it. Window mode focuses on just the selected milestone pair.</div>
    """, unsafe_allow_html=True)
with jh_right:
    rr1, rr2 = st.columns(2)
    with rr1:
        agg_mode = st.radio("Aggregation", ["Earliest", "Latest"], horizontal=True, label_visibility="collapsed", key="agg_mode")
        st.caption("Aggregation")
    with rr2:
        journey_mode = st.radio("Mode", ["Whole Journey", "Window"], horizontal=True, label_visibility="collapsed", key="journey_mode")
        st.caption("Mode")

whole_journey = (journey_mode == "Whole Journey")

# Recompute shipment_lt if aggregation changed (only Earliest/Latest matters now)
if agg_mode != "Earliest":
    shipment_lt = compute_shipment_leadtimes(raw_df, start_ms, end_ms, agg_mode, True)
    shipment_lt = apply_top_n_lanes_filter(shipment_lt, top_n_lanes)

# Journey diagram
st.markdown(render_journey_diagram(start_ms, end_ms, whole_journey), unsafe_allow_html=True)

# Compute scope metrics for all segments + whole journey
scope_stats = {}
if not shipment_lt.empty:
    j_col = _journey_col_name()
    if j_col in shipment_lt.columns:
        s = shipment_lt[j_col].dropna()
        if not s.empty:
            med_h = float(s.median())
            pxx_h = _safe_quantile(s, percentile_p / 100.0)
            mad_h = float((s - med_h).abs().median())
            scope_stats["whole"] = {"name": f"Whole Journey · {start_ms} → {end_ms}" if not whole_journey else "Whole Journey · CEP → CER",
                                    "median_d": round(med_h / 24, 1), "pxx_d": round(pxx_h / 24, 1) if pxx_h else None,
                                    "mad_d": round(mad_h / 24, 1), "ships": eligible, "col": j_col}
    if whole_journey:
        for a, b in SEGMENTS:
            col = _segment_col_name(a, b)
            if col in shipment_lt.columns:
                s = shipment_lt[col].dropna()
                if not s.empty:
                    med_h = float(s.median())
                    pxx_h = _safe_quantile(s, percentile_p / 100.0)
                    mad_h = float((s - med_h).abs().median())
                    scope_stats[f"{a}-{b}"] = {"name": f"{a} → {b} segment",
                                              "median_d": round(med_h / 24, 1),
                                              "pxx_d": round(pxx_h / 24, 1) if pxx_h else None,
                                              "mad_d": round(mad_h / 24, 1),
                                              "ships": int(shipment_lt[col].notna().sum()), "col": col}

# Bottleneck detection
seg_stats_list = []
for _a, _b in SEGMENTS:
    _key = f"{_a}-{_b}"
    if _key in scope_stats:
        _item = {"segment_key": _key, "start_ms": _a, "end_ms": _b}
        _item.update(scope_stats[_key])
        seg_stats_list.append(_item)
bottleneck = compute_bottleneck(seg_stats_list) if whole_journey else None
bottleneck_key = bottleneck["segment_key"] if bottleneck else None

# Whole Journey card
whole_stat = scope_stats.get("whole", {})
wj_active = st.session_state["active_scope"] == "whole"

if whole_journey:
    st.markdown(render_whole_journey_card(wj_active,
                                          whole_stat.get("median_d"), whole_stat.get("pxx_d"),
                                          whole_stat.get("mad_d"), whole_stat.get("ships", 0)),
                unsafe_allow_html=True)
    if st.button("Select Whole Journey scope", key="select_whole", use_container_width=False):
        st.session_state["active_scope"] = "whole"; st.rerun()

    # 7 segment cards as a row of styled buttons
    st.markdown('<div class="v7-segment-grid-wrap">', unsafe_allow_html=True)
    seg_cols = st.columns(7)
    for i, (a, b) in enumerate(SEGMENTS):
        key = f"{a}-{b}"
        st_data = scope_stats.get(key, {})
        med = st_data.get("median_d", "—"); pxx = st_data.get("pxx_d", "—")
        is_active = st.session_state["active_scope"] == key
        is_bottleneck = bottleneck_key == key
        # Pick the button key prefix so CSS targets the active/bottleneck state
        prefix = "seg_active_" if is_active else ("seg_bottleneck_" if is_bottleneck else "seg_")
        label = f"{a} → {b}\nMedian: {med}d\nP{percentile_p}: {pxx}d"
        if is_bottleneck:
            label = "⚠ " + label
        with seg_cols[i]:
            if st.button(label, key=f"{prefix}{key}", use_container_width=True):
                st.session_state["active_scope"] = key; st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

# Active Scope summary bar
active_key = st.session_state["active_scope"]
if not whole_journey:
    active_key = "whole"  # In window mode, scope is always the configured window
active_stat = scope_stats.get(active_key, {})
st.markdown(f"""
<div class="v7-jh-summary">
  <div class="v7-jws-block"><div class="l">Active scope</div><div class="v" style="font-size: 14px;">{active_stat.get('name', '—')}</div></div>
  <div class="v7-jws-block"><div class="l">Scope Median</div><div class="v">{active_stat.get('median_d', '—')}<span class="u">d</span></div></div>
  <div class="v7-jws-block"><div class="l">Scope P{percentile_p}</div><div class="v">{active_stat.get('pxx_d', '—')}<span class="u">d</span></div></div>
  <div class="v7-jws-block"><div class="l">Scope MAD</div><div class="v">{active_stat.get('mad_d', '—')}<span class="u">d</span></div></div>
  <div class="v7-jws-block"><div class="l">Eligible Shipments</div><div class="v">{active_stat.get('ships', 0):,}</div></div>
</div>
</div>
""", unsafe_allow_html=True)

# ---- Bottleneck callout ----
if bottleneck and whole_journey:
    a = bottleneck.get("start_ms")
    b = bottleneck.get("end_ms")
    bk = bottleneck.get("segment_key")
    if a and b and bk:
        st.markdown(f"""
        <div class="v7-bottleneck-callout">
          <div class="v7-bc-icon">🚧</div>
          <div class="v7-bc-text">
            <h3>Bottleneck: <strong>{a} → {b}</strong> ({MILESTONE_LONG[a][0]} → {MILESTONE_LONG[b][0]})</h3>
            <p>This segment has the highest P{percentile_p}/median ratio ({bottleneck['ratio']:.2f}×). Median {bottleneck['median_d']}d but worst 20%+ take {bottleneck['pxx_d']}d. Improving this segment alone would tighten end-to-end P{percentile_p} by ~{bottleneck['improvement_d']}d.</p>
          </div>
        </div>
        """, unsafe_allow_html=True)
        if st.button("Drill into this segment →", key="drill_bottleneck"):
            st.session_state["active_scope"] = bk; st.rerun()

# ---- Lane Focus selector ----
all_lanes = lane_counts["Lane"].tolist() if not lane_counts.empty else []
if not all_lanes:
    st.warning("No lanes available after filtering.")
    st.stop()

st.markdown('<div class="v7-lane-context">', unsafe_allow_html=True)
lc1, lc2 = st.columns([2, 5])
with lc1:
    focus_lane = st.selectbox("Focus lane (insights scope)", all_lanes, label_visibility="collapsed")
st.markdown('</div>', unsafe_allow_html=True)

# ---- Compute insights for the focus lane ----
duration_configs = build_duration_configs(start_ms, end_ms, whole_journey)
journey_cfg = duration_configs[0]  # the journey window config

lane_summary, carrier_recs = compute_insights_for_metric(
    shipment_lt, journey_cfg, int(percentile_p),
    pct_threshold_on=False, pct_threshold_pct=0.0,
    rec_threshold_on=False, rec_threshold_pct=0.0,
)

lane_row = lane_summary[lane_summary["LANE"] == focus_lane]
lane_carriers = carrier_recs[(carrier_recs["LANE"] == focus_lane) & (carrier_recs["RECOMMENDATION_ELIGIBLE"])]
lane_carriers = lane_carriers.sort_values("RANK_IN_LANE", na_position="last")

lane_median_d = lane_row["LANE_MEDIAN_D"].iloc[0] if not lane_row.empty else None
lane_median_h = lane_row["LANE_MEDIAN_H"].iloc[0] if not lane_row.empty else None
lane_pxx_d = lane_row["LANE_PXX_D"].iloc[0] if not lane_row.empty else None
lane_pxx_h = lane_row["LANE_PXX_H"].iloc[0] if not lane_row.empty else None

# ---- Insight cards row 1: Best Carrier | Top 5 | Problematic ----
st.markdown(f"""
<div class="v7-section-head">
  <div>
    <div class="v7-section-eyebrow">Focus lane only</div>
    <div class="v7-section-title">Insights for {focus_lane}</div>
    <div class="v7-section-subtitle">Scoped to: <strong style="color: var(--text);">{active_stat.get('name', '—')}</strong></div>
  </div>
</div>
""", unsafe_allow_html=True)

i1, i2, i3 = st.columns(3)
with i1:
    best = lane_carriers.iloc[0].to_dict() if not lane_carriers.empty else None
    st.markdown(render_best_carrier(best, lane_median_d), unsafe_allow_html=True)
with i2:
    st.markdown(render_top5(lane_carriers, lane_median_d), unsafe_allow_html=True)
with i3:
    # Compute dynamic problematic
    scope_col = active_stat.get("col", _journey_col_name())
    prob_mode = st.session_state["prob_mode"]
    prob_df = compute_problematic_shipments(shipment_lt, focus_lane, scope_col,
                                            lane_median_h, lane_pxx_h, prob_mode)
    # Compute trimmed median (top 5%)
    skew = compute_skew_analysis(shipment_lt[shipment_lt["LANE"] == focus_lane][scope_col],
                                 lane_median_d) if scope_col in shipment_lt.columns else []
    trimmed = skew[0]["new_median_d"] if skew else lane_median_d
    st.markdown(render_problematic_card(active_stat.get("name", "—"), prob_df, prob_mode, lane_median_d, trimmed),
                unsafe_allow_html=True)
    # Problematic controls
    pm1, pm2 = st.columns([3, 1])
    with pm1:
        new_mode = st.radio("Problematic mode", ["Above Median", "Above P-XX"], horizontal=True,
                            label_visibility="collapsed", index=0 if prob_mode == "Above Median" else 1, key="prob_mode_radio")
        if new_mode != prob_mode:
            st.session_state["prob_mode"] = new_mode; st.rerun()
    with pm2:
        if not prob_df.empty:
            prob_excel = write_problematic_excel(prob_df, active_stat.get("name", "Scope"), prob_mode)
            st.download_button("↓ Export", prob_excel, file_name="problematic_shipments.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_prob", use_container_width=True)

# ---- Insight cards row 2: Skew | Trend ----
st.markdown("<div style='height: 14px'></div>", unsafe_allow_html=True)
s1, s2 = st.columns([1, 1.2])
with s1:
    st.markdown(render_skew_card(skew, lane_median_d), unsafe_allow_html=True)
with s2:
    st.markdown('<div class="v7-insight-card">', unsafe_allow_html=True)
    th1, th2 = st.columns([3, 1])
    with th1:
        st.markdown('<div class="v7-ic-title">Lead Time Trend <span class="v7-tag-ai">Derived</span></div>', unsafe_allow_html=True)
    with th2:
        granularity = st.radio("Granularity", ["Month", "Quarter"], horizontal=True, label_visibility="collapsed", key="trend_gran")
    trend_data = compute_trend_buckets(shipment_lt, focus_lane, granularity)
    if len(trend_data) < 2:
        st.markdown('<p style="color: var(--text-tertiary); font-size: 12.5px; padding: 20px 0;">Need ≥ 2 buckets to display trend. Try switching to Month granularity or upload a file with a wider date range.</p>', unsafe_allow_html=True)
    else:
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=[d["bucket"] for d in trend_data], y=[d["median_d"] for d in trend_data],
                                 mode="lines+markers", line=dict(color="#818cf8", width=2.5),
                                 marker=dict(color="#ec4899", size=8), fill="tozeroy",
                                 fillcolor="rgba(99,102,241,0.15)"))
        fig.update_layout(height=180, margin=dict(l=10, r=10, t=10, b=30),
                          paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                          font=dict(color="#9ca3af", size=11), showlegend=False,
                          xaxis=dict(gridcolor="rgba(255,255,255,0.04)"),
                          yaxis=dict(gridcolor="rgba(255,255,255,0.04)"))
        st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})
    st.markdown('<div style="margin-top: 8px; padding-top: 8px; border-top: 1px solid var(--border); font-size: 11.5px; color: var(--text-tertiary);">Bucketing uses <code style="font-family: SF Mono, monospace; color: var(--accent-light); background: var(--bg-elev); padding: 1px 4px; border-radius: 3px;">SHIPMENT_MONTH_YEAR</code> already derived by the tool.</div></div>', unsafe_allow_html=True)

# ---- Insight cards row 3: Reliability | Concentration ----
st.markdown("<div style='height: 14px'></div>", unsafe_allow_html=True)
r1, r2 = st.columns(2)
with r1:
    # Compute reliability for focus lane vs all lanes
    all_medians = lane_summary["LANE_MEDIAN_D"].dropna().tolist() if not lane_summary.empty else []
    # Compute lane-level MAD per lane (simple: median of |shipment_lead - lane_median| in each lane)
    lane_mads = {}
    for ln in lane_summary["LANE"].tolist():
        sub = shipment_lt[shipment_lt["LANE"] == ln][_journey_col_name()].dropna()
        if not sub.empty:
            med = float(sub.median())
            lane_mads[ln] = float((sub - med).abs().median()) / 24.0
    focus_mad = lane_mads.get(focus_lane)
    all_mads = list(lane_mads.values())
    score_info = compute_reliability_score(lane_median_d, focus_mad, all_medians, all_mads)
    # Rank
    lane_summary_with_score = lane_summary.copy()
    lane_summary_with_score["_SCORE"] = lane_summary_with_score.apply(
        lambda r: compute_reliability_score(r["LANE_MEDIAN_D"], lane_mads.get(r["LANE"]), all_medians, all_mads)["score"], axis=1)
    rank_series = lane_summary_with_score.sort_values("_SCORE", ascending=False).reset_index(drop=True)
    rank = (rank_series.index[rank_series["LANE"] == focus_lane].tolist() or [0])[0] + 1
    n_total = len(lane_summary_with_score)
    st.markdown(render_reliability_card(score_info, n_total, rank), unsafe_allow_html=True)

with r2:
    # Concentration card
    st.markdown('<div class="v7-insight-card">', unsafe_allow_html=True)
    st.markdown('<div class="v7-ic-head"><div class="v7-ic-title">Carrier Concentration <span class="v7-tag-ai">Classification</span></div></div>', unsafe_allow_html=True)
    cc_left, cc_right = st.columns([1, 1.3])
    with cc_left:
        if not lane_carriers.empty:
            shares = lane_carriers.head(8)[["CARRIER_NAME", "CARRIER_SHARE_PCT"]].dropna()
            fig = go.Figure(data=[go.Pie(labels=shares["CARRIER_NAME"].tolist(),
                                         values=shares["CARRIER_SHARE_PCT"].tolist(),
                                         hole=0.65, marker=dict(colors=["#6366f1", "#ec4899", "#10b981", "#3b82f6",
                                                                         "#f97316", "#fbbf24", "#9ca3af", "#a78bfa"]),
                                         textinfo="none")])
            fig.update_layout(height=180, margin=dict(l=0, r=0, t=0, b=0),
                              paper_bgcolor="rgba(0,0,0,0)", showlegend=False)
            st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})
    with cc_right:
        if not lane_carriers.empty:
            shares = lane_carriers.head(8)[["CARRIER_NAME", "CARRIER_SHARE_PCT"]].dropna()
            top_share = shares["CARRIER_SHARE_PCT"].iloc[0] if not shares.empty else 0
            top_two = shares["CARRIER_SHARE_PCT"].head(2).sum() if len(shares) >= 2 else top_share
            classification = classify_concentration(top_share, top_two)
            for _, row in shares.iterrows():
                st.markdown(f'<div style="display: flex; justify-content: space-between; font-size: 11.5px; padding: 3px 0; color: var(--text-secondary);"><span>{row["CARRIER_NAME"]}</span><strong style="color: var(--text);">{row["CARRIER_SHARE_PCT"]}%</strong></div>', unsafe_allow_html=True)
            st.markdown(f'<div class="v7-conc-flag {classification["color"]}">⚠️ {classification["label"]}</div>', unsafe_allow_html=True)
            st.markdown('<div style="margin-top: 10px; padding: 8px 12px; background: var(--bg-elev); border-radius: 6px; font-size: 11px; color: var(--text-tertiary); line-height: 1.5;"><strong style="color: var(--text-secondary);">Classification:</strong> top carrier &lt; 40% = Low, 40–60% = Moderate, &gt; 60% = High.</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

# ---- Carrier Lane Lead Detail (Network-wide) ----
st.markdown(f"""
<div class="v7-section-head">
  <div>
    <div class="v7-section-eyebrow">Network-wide</div>
    <div class="v7-section-title">Carrier Lane Lead Detail · All Lanes</div>
    <div class="v7-section-subtitle">Every lane × every carrier · matches the main file your existing Streamlit exports today</div>
  </div>
</div>
""", unsafe_allow_html=True)

report_df = build_carrier_lane_report(shipment_lt, int(percentile_p), True, 0, duration_configs)
st.markdown(render_analyst_table(report_df, int(percentile_p)), unsafe_allow_html=True)

main_excel = write_excel_final(raw_df, report_df, duration_configs, int(percentile_p))
st.download_button("↓ Download Carrier Lane Lead Detail (Excel)",
                   main_excel, file_name=f"carrier_lane_lead_{start_ms}_to_{end_ms}.xlsx",
                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_main")

st.markdown("<div style='height: 40px'></div>", unsafe_allow_html=True)
st.markdown('<div style="text-align: center; font-size: 11px; color: var(--text-quaternary);">Lead Time Optimization · Ocean · v7 Streamlit · project44</div>', unsafe_allow_html=True)
